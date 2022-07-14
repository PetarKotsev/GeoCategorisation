# Streamlit UI
from random import choices
import streamlit as st 
# Keras load model
from keras.models import load_model
# Excell
import openpyxl as xl
import numpy as np
# Collections
import collections

model = load_model('./models/newest.model')

st.sidebar.title('Selections')

# load data for the selecets
## loading an existing workbook
wb = xl.load_workbook('Geology Input for Petar Updated April 11th.xlsx')

## Select the worksheet
categorical_items_sheet = wb['Sheet3']
## Select the rectangle where the data is situated
categorical_items = list(categorical_items_sheet.iter_cols(min_row=1, max_row=24, min_col=1, max_col=6,  values_only=True))
## get rid of all none-valued cells
categorical_items = [tuple(xi for xi in x if xi is not None) for x in categorical_items]

## Grab all feature categories from sheet 4
item_categories_sheet = wb['Sheet4']
terain_property_classes = tuple(item_categories_sheet.iter_rows(min_row=1, max_row=1, min_col=1, max_col=6,  values_only=True))[0]

## Grab all unique terraign properties
props_sheet = wb['Sheet1']
terain_properties = list(props_sheet.iter_cols(min_row=1, max_col=1, min_col=1, values_only=True))[0]
terain_properties = list(map(lambda x: x.strip(), terain_properties))

## Grab all types of terraign and their discriptions
categories_sheet = wb['Sheet5']
categories = list(categories_sheet.iter_cols(min_row=1, max_row=10, min_col=1, max_col=1,  values_only=True))[0]
category_texts = list(categories_sheet.iter_cols(min_row=1, max_row=10, min_col=2, max_col=2,  values_only=True))[0]

## Handle the changes in the ui
def changeHandler():
    l = {}
    # accumolate all selections
    choices = st.session_state.choices + st.session_state.mult_choices
    # get a prediction
    predictions = getPredictions(choices, terain_properties)[0]
    # put all predictions into a dict
    for inx, clas in enumerate(categories):
        l[clas] = predictions[inx]
    # order the items of the dict
    new_l = {k: v for k, v in sorted(l.items(), key=lambda item: item[1], reverse=True)}
    # get the top prediction
    top_selection = list(new_l.keys())[0]
    # get the index of the top selection
    selection_inx = categories.index(top_selection)
    
    ## make string
    string = category_texts[selection_inx].replace("[user's input strings]", str(choices)).replace("[probability of prediction]", str(new_l[top_selection]))
    ## Write to screen
    st.write(string)
    st.write(new_l)

def getPredictions(predictionArr, properties):
    # prepare the statesents for input
    statements = np.zeros((1, len(properties)))
    for line in statements:
        for prop in predictionArr:
            line[properties.index(prop)] = 1.0
    # return the prediction
    return model.predict(statements)

## All choises
st.session_state.choices = []

## for each terraign property class make a selector
for inx, clas in enumerate(terain_property_classes):
    st.session_state.choices.append(st.sidebar.selectbox(clas, categorical_items[inx], key=clas))
    ## when the first 3 classes are appended add the rest to one multiselect (they are all specieal features from inx 2 forward)
    if inx >=2:
        st.session_state.mult_choices = st.sidebar.multiselect('Special features', np.concatenate(categorical_items[3:]))
        break

## add a button that would make a prediction
st.sidebar.button('Update', on_click=changeHandler)




